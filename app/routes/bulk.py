import io
import json
import re
from datetime import datetime, timedelta

import pandas as pd
from flask import Blueprint, render_template, request, send_file, session

from app.routes.auth import login_required
from app.services.q360 import Q360Service, CATEGORIES, COMPANIES
from app.db import get_db

bp = Blueprint('bulk', __name__, url_prefix='/bulk')

DAY_COLS = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']
DAY_OFFSETS = {'Mon': 0, 'Tue': 1, 'Wed': 2, 'Thu': 3, 'Fri': 4, 'Sat': 5, 'Sun': 6}
DEFAULT_CATEGORY = 'VOICE- CCAPPDEV'
DEFAULT_COMPANY = 'CONNEX TELECOMMUNICATIONS INC.'


def _svc():
    return Q360Service(session['user_id'], session['password'])


def _fmt_time(dt):
    """Format datetime as HH:MM string."""
    return dt.strftime('%H:%M')


def _week_monday(excel_week):
    """Convert Excel week (0-51) to its Monday. Week 0 = Mon of ISO week 52 of previous year."""
    year = datetime.now().year
    week_0_monday = datetime.fromisocalendar(year - 1, 52, 1)
    return week_0_monday + timedelta(weeks=int(excel_week))


def _normalize(s):
    return s.lower().strip() if s else ''


def _build_recommendations(task_rows):
    """For task rows missing Q360ID, recommend from earlier weeks by customer+project similarity."""
    from collections import defaultdict

    by_cust_proj = defaultdict(lambda: defaultdict(set))
    by_cust      = defaultdict(lambda: defaultdict(set))
    by_user      = defaultdict(lambda: defaultdict(set))
    # Track best (customer, project) seen for each (username, q360id) — most recent week wins
    q360_details = {}  # (username, q360id) -> (customer, project, week)

    for r in task_rows:
        if not r['needs_attention'] and r['q360id']:
            u, ck, pk, wk = r['username'], _normalize(r['customer']), _normalize(r['project']), r['week_num']
            by_cust_proj[(u, ck, pk)][wk].add(r['q360id'])
            by_cust[(u, ck)][wk].add(r['q360id'])
            by_user[u][wk].add(r['q360id'])
            if r['customer']:  # only index rows with actual project data
                key = (u, r['q360id'])
                if key not in q360_details or q360_details[key][2] < wk:
                    q360_details[key] = (r['customer'], r['project'], wk)

    def _collect(mapping, key, max_week):
        ids = set()
        for wk, qids in mapping.get(key, {}).items():
            if wk < max_week:
                ids.update(qids)
        return ids

    for r in task_rows:
        if not r['needs_attention']:
            r['recommended_q360ids'] = []
            r['suggested'] = False
            continue
        u, ck, pk, wk = r['username'], _normalize(r['customer']), _normalize(r['project']), r['week_num']
        candidates = _collect(by_cust_proj, (u, ck, pk), wk)
        if not candidates:
            candidates = _collect(by_cust, (u, ck), wk)
        if not candidates:
            candidates = _collect(by_user, u, wk)
        r['recommended_q360ids'] = sorted(candidates)
        r['suggested'] = False



def _assign_time_slots(task_rows):
    """Assign non-overlapping start/end times per (username, day), ordered by row_num."""
    from collections import defaultdict
    # Build per (username, date) ordered list of (row, day) with hours
    slots = defaultdict(list)
    for r in task_rows:
        for day, d in r['days'].items():
            if d['hours'] > 0:
                slots[(r['username'], d['date'])].append((r['row_num'], r, day))

    for entries in slots.values():
        entries.sort(key=lambda x: x[0])
        current = datetime.strptime(entries[0][1]['days'][entries[0][2]]['date'] + ' 08:00',
                                    '%Y-%m-%d %H:%M')
        for _, r, day in entries:
            d = r['days'][day]
            d['start_time'] = _fmt_time(current)
            current_end = current + timedelta(hours=d['hours'])
            d['end_time'] = _fmt_time(current_end)
            current = current_end


def _fill_missing_weeks(task_rows):
    """For weeks present for some users but not others, synthesize rows from nearest week."""
    from collections import defaultdict

    all_weeks = sorted(set(r['week_num'] for r in task_rows))
    if len(all_weeks) <= 1:
        return []

    by_user = defaultdict(list)
    for r in task_rows:
        by_user[r['username']].append(r)

    max_row_num = max(r['row_num'] for r in task_rows)
    new_rows = []

    for username, rows in by_user.items():
        user_weeks = sorted(set(r['week_num'] for r in rows))
        for week_num in all_weeks:
            if week_num in user_weeks:
                continue

            # Find closest week (prefer previous, fall back to next)
            prev = [w for w in user_weeks if w < week_num]
            nxt  = [w for w in user_weeks if w > week_num]
            source_week = prev[-1] if prev else (nxt[0] if nxt else None)
            if source_week is None:
                continue

            source_rows = [r for r in rows if r['week_num'] == source_week]
            week_start = _week_monday(week_num)
            week_end   = week_start + timedelta(days=4)
            week_range = (f"Wk {week_num}: "
                          f"{week_start.strftime('%b')} {week_start.day}–{week_end.day}, {week_start.year}")

            for src in source_rows:
                max_row_num += 1
                new_days = {}
                for day in DAY_COLS:
                    date = (week_start + timedelta(days=DAY_OFFSETS[day])).strftime('%Y-%m-%d')
                    new_days[day] = {
                        'hours': src['days'][day]['hours'],
                        'date': date,
                        'start_time': '',
                        'end_time': '',
                    }
                new_rows.append({
                    'row_num':            max_row_num,
                    'week_num':           week_num,
                    'week_range':         week_range,
                    'username':           username,
                    'employee':           src['employee'],
                    'customer':           src['customer'],
                    'project':            src['project'],
                    'q360id':             src['q360id'],
                    'needs_attention':    src['needs_attention'],
                    'suggested':          src['suggested'],
                    'category':           src['category'],
                    'company':            src['company'],
                    'comment':            src['comment'],
                    'days':               new_days,
                    'recommended_q360ids': list(src.get('recommended_q360ids', [])),
                    'missing_week':       True,
                    'missing_guessed_from': source_week,
                })
    return new_rows


def _resolve_usernames_from_team(grouped):
    """For employees whose name matches a team_member row, use that DB username.
    Runs before manual corrections so username_map entries still take priority."""
    db = get_db()
    rows = db.execute(
        'SELECT username, name FROM team_member WHERE name IS NOT NULL AND name != ""'
    ).fetchall()
    name_to_username = {r['name'].strip().lower(): r['username'] for r in rows}
    for u in grouped:
        employee = (u.get('employee') or '').strip()
        if not employee:
            continue
        db_username = name_to_username.get(employee.lower())
        if db_username:
            u['username'] = db_username
            for r in u['rows']:
                r['username'] = db_username
    return grouped


def _merge_duplicate_usernames(grouped):
    """Merge groups that share the same username (can happen after username resolution)."""
    seen = {}
    merged = []
    for u in grouped:
        un = u['username']
        if un in seen:
            seen[un]['rows'].extend(u['rows'])
        else:
            seen[un] = u
            merged.append(u)
    return merged


def _apply_username_corrections(grouped):
    """Replace Excel-derived usernames with saved Q360 corrections from the DB."""
    rows = get_db().execute('SELECT employee_name, q360_username FROM username_map').fetchall()
    corrections = {r['employee_name']: r['q360_username'] for r in rows}
    if not corrections:
        return grouped
    for u in grouped:
        corrected = corrections.get(u['employee'])
        if corrected:
            u['username'] = corrected
            for r in u['rows']:
                r['username'] = corrected
    return grouped


def _filter_by_date(grouped, mode):
    """Filter grouped rows to only those whose week overlaps the selected date range.
    A week overlaps a range when: week_monday <= range_end AND week_sunday >= range_start.
    Full weeks are always kept intact even if they straddle a month/year boundary.
    """
    if mode == 'all':
        return grouped
    import calendar as _cal
    today = datetime.today().date()

    if mode == 'week':
        range_start = today - timedelta(days=today.weekday())
        range_end   = range_start + timedelta(days=6)
    elif mode == 'month':
        range_start = today.replace(day=1)
        range_end   = today.replace(day=_cal.monthrange(today.year, today.month)[1])
    elif mode == 'last_month':
        first_of_this = today.replace(day=1)
        range_end   = first_of_this - timedelta(days=1)
        range_start = range_end.replace(day=1)
    elif mode == 'year':
        range_start = today.replace(month=1, day=1)
        range_end   = today.replace(month=12, day=31)
    else:
        return grouped

    result = []
    for u in grouped:
        filtered = []
        for r in u['rows']:
            monday = _week_monday(r['week_num']).date()
            sunday = monday + timedelta(days=6)
            # Overlap: week touches the range at all
            if monday <= range_end and sunday >= range_start:
                filtered.append(r)
        if filtered:
            result.append({**u, 'rows': filtered})
    return result


def _parse_excel(file):
    """Parse uploaded Excel file. Returns list of user dicts, each with task rows."""
    filename = file.filename.lower()
    if filename.endswith('.xlsb'):
        df = pd.read_excel(file, engine='pyxlsb', sheet_name=0)
    else:
        df = pd.read_excel(file, engine='openpyxl', sheet_name=0)

    from collections import defaultdict

    task_rows = []
    row_num = 0

    for _, row in df.iterrows():
        username = str(row.get('Username', '') or '').strip()
        employee = str(row.get('Employee', '') or '').strip()
        if not username or username == 'nan':
            # Derive username from Employee full name: first letter + last name
            if employee and employee != 'nan':
                parts = employee.strip().split()
                if len(parts) >= 2:
                    username = (parts[0][0] + parts[-1]).upper()
                else:
                    continue
            else:
                continue

        week_val = row.get('Week')
        if week_val is None or str(week_val) == 'nan':
            continue
        week_num = int(float(week_val))
        week_start = _week_monday(week_num)

        def _clean(v):
            s = str(v or '').strip()
            return '' if s == 'nan' else s

        customer = re.sub(r'^\d+\.', '', _clean(row.get('Customer', ''))).strip()
        project  = _clean(row.get('Project', ''))
        comment  = _clean(row.get('Comment', ''))


        # Build per-day data (all days, 0 if no hours)
        days = {}
        has_any = False
        for day in DAY_COLS:
            h = row.get(day)
            hours = 0.0 if (h is None or str(h) == 'nan') else float(h)
            date = (week_start + timedelta(days=DAY_OFFSETS[day])).strftime('%Y-%m-%d')
            days[day] = {'hours': hours, 'date': date, 'start_time': '', 'end_time': ''}
            if hours > 0:
                has_any = True

        if not has_any:
            continue

        week_end = week_start + timedelta(days=4)
        week_range = (f"Wk {week_num}: "
                      f"{week_start.strftime('%b')} {week_start.day}–{week_end.day}, {week_start.year}")

        row_num += 1
        task_rows.append({
            'row_num': row_num,
            'week_num': week_num,
            'week_range': week_range,
            'username': username,
            'employee': employee,
            'customer': customer,
            'project': project,
            'q360id': '',
            'needs_attention': True,
            'suggested': False,
            'category': '',
            'company': DEFAULT_COMPANY,
            'comment': comment,
            'days': days,
            'recommended_q360ids': [],
            'missing_week': False,
            'missing_guessed_from': None,
            'project_guessed': False,
        })

    _build_recommendations(task_rows)

    # Fill missing weeks before assigning time slots so slots cover synthetic rows too
    missing = _fill_missing_weeks(task_rows)
    task_rows.extend(missing)
    task_rows.sort(key=lambda r: (r['username'], r['week_num'], r['row_num']))

    _assign_time_slots(task_rows)

    # Group by user, preserving order of first appearance
    seen_users = []
    by_user_map = {}
    for r in task_rows:
        u = r['username']
        if u not in by_user_map:
            seen_users.append(u)
            by_user_map[u] = []
        by_user_map[u].append(r)

    return [{'username': u, 'employee': by_user_map[u][0]['employee'], 'rows': by_user_map[u]}
            for u in seen_users]


@bp.route('/')
@login_required
def index():
    return render_template('bulk/index.html')


@bp.route('/parse', methods=['POST'])
@login_required
def parse():
    import json as _json
    import traceback as _tb
    from flask import Response, stream_with_context

    file = request.files.get('file')
    if not file:
        return '<div class="alert alert-danger">Please provide a file.</div>'
    date_filter = request.form.get('date_filter', 'month')
    # Read file bytes eagerly so the stream can start immediately (parsing happens inside the generator)
    import io as _io
    file_bytes = file.read()
    file_name  = file.filename

    def _evt(obj):
        return f'data: {_json.dumps(obj)}\n\n'

    def generate():
        from concurrent.futures import ThreadPoolExecutor, as_completed
        # Parse Excel inside the stream so the SSE connection opens immediately
        yield _evt({'type': 'progress', 'done': 0, 'total': 1, 'user': '', 'msg': 'Reading Excel file\u2026'})
        try:
            fake_file = _io.BytesIO(file_bytes)
            fake_file.filename = file_name
            grouped = _parse_excel(fake_file)
            grouped = _resolve_usernames_from_team(grouped)
            grouped = _apply_username_corrections(grouped)
            grouped = _merge_duplicate_usernames(grouped)
            grouped = _filter_by_date(grouped, date_filter)
        except Exception as e:
            yield _evt({'type': 'error', 'html':
                        f'<div class="alert alert-danger">Failed to parse file: {e}</div>'})
            return
        if not grouped:
            yield _evt({'type': 'error', 'html':
                        '<div class="alert alert-warning">No entries found for the selected date range.</div>'})
            return

        try:
            svc = _svc()
        except Exception as e:
            yield _evt({'type': 'error', 'html':
                        f'<div class="alert alert-danger">Q360 login failed: {e}</div>'})
            return

        try:
            usernames = list({u['username'] for u in grouped})
            total = len(usernames)
            live_by_user = {}

            # Build name lookup: username → "Full Name (username)" or just username
            db = get_db()
            name_rows = db.execute('SELECT username, name FROM team_member').fetchall()
            name_map = {r['username']: r['name'] for r in name_rows if r['name']}
        except Exception as e:
            yield _evt({'type': 'error', 'html':
                        f'<div class="alert alert-danger">Initialisation error: {_tb.format_exc()}</div>'})
            return

        def _display_user(un):
            n = name_map.get(un, '')
            return f'{n} ({un})' if n else un

        # Load cached project lists (fresh = fetched within last 24 hours)
        cache_rows = db.execute(
            "SELECT username, projects_json FROM project_cache "
            "WHERE fetched_at >= datetime('now', '-24 hours', 'localtime')"
        ).fetchall()
        cached = {r['username']: json.loads(r['projects_json']) for r in cache_rows}

        to_fetch = [un for un in usernames if un not in cached]
        # Users served from cache count as already done
        for un in usernames:
            if un in cached:
                live_by_user[un] = cached[un]

        if to_fetch:
            yield _evt({'type': 'progress', 'done': len([un for un in usernames if un in cached]),
                        'total': total, 'user': '', 'msg': 'Connecting to Q360\u2026'})

            def _fetch(username):
                try:
                    return username, svc.get_projects(username)
                except Exception:
                    return username, {}

            with ThreadPoolExecutor(max_workers=min(len(to_fetch), 10)) as pool:
                futures = {pool.submit(_fetch, un): un for un in to_fetch}
                for f in as_completed(futures):
                    un, projects = f.result()
                    live_by_user[un] = projects
                    # Persist to cache
                    db.execute(
                        "INSERT INTO project_cache (username, projects_json, fetched_at) "
                        "VALUES (?, ?, datetime('now','localtime')) "
                        "ON CONFLICT(username) DO UPDATE SET "
                        "projects_json=excluded.projects_json, fetched_at=excluded.fetched_at",
                        (un, json.dumps(projects))
                    )
                    db.commit()
                    yield _evt({'type': 'progress', 'done': len(live_by_user),
                                'total': total, 'user': _display_user(un),
                                'msg': f'Fetched projects for {un}'})
        else:
            yield _evt({'type': 'progress', 'done': total, 'total': total,
                        'user': '', 'msg': 'Loaded from cache'})

        # Upsert team members: add any new usernames from the upload
        try:
            admin_row = db.execute(
                'SELECT team FROM team_member WHERE username = ?', (session['user_id'],)
            ).fetchone()
            default_team = 'All'
            import re as _re
            for u in grouped:
                username = u['username']
                # Skip blank, too-short, or obviously-derived garbage usernames
                # Valid Q360 usernames have ≥4 chars and contain at least one letter
                if not username or len(username) < 4:
                    continue
                if not _re.search(r'[a-zA-Z]', username):
                    continue
                employee_name = u.get('employee', '') or ''
                existing = db.execute(
                    'SELECT id FROM team_member WHERE username = ? OR '
                    '(name != "" AND LOWER(name) = LOWER(?))',
                    (username, employee_name)
                ).fetchone()
                if not existing:
                    db.execute(
                        'INSERT INTO team_member (username, name, team) VALUES (?, ?, ?)',
                        (username, employee_name, default_team)
                    )
                elif employee_name:
                    # Update name if we now have one and it was blank before
                    db.execute(
                        'UPDATE team_member SET name = ? WHERE username = ? AND (name IS NULL OR name = ?)',
                        (employee_name, username, '')
                    )
            db.commit()
        except Exception:
            pass

        try:
            html = _do_parse(grouped, live_by_user, date_filter)
            yield _evt({'type': 'complete', 'html': html})
        except Exception:
            err = _tb.format_exc()
            yield _evt({'type': 'error', 'html':
                        f'<div class="alert alert-danger"><strong>Unexpected error:</strong>'
                        f'<pre style="font-size:.75rem;white-space:pre-wrap">{err}</pre></div>'})

    return Response(
        stream_with_context(generate()),
        mimetype='text/event-stream',
        headers={'X-Accel-Buffering': 'no', 'Cache-Control': 'no-cache'},
    )


def _do_parse(grouped, live_by_user, date_filter='month'):
    # Determine which day columns have any data
    active_days = [d for d in DAY_COLS if any(
        r['days'][d]['hours'] > 0 for u in grouped for r in u['rows']
    )]

    def _match_score(excel_customer, excel_project, q360_desc):
        search = ' '.join(filter(None, [excel_customer, excel_project])).lower()
        desc = q360_desc.lower()
        search = re.sub(r'[^\w\s]', ' ', search)
        desc   = re.sub(r'[^\w\s]', ' ', desc)
        words = [w for w in search.split() if len(w) > 2]
        if not words:
            return 0.0
        return sum(1 for w in words if w in desc) / len(words)

    # Non-billable Q360 description prefixes
    _NON_BILL_PREFIXES = (
        'ADMINISTRATION', 'ON-CALL', 'PERSONAL TIME-OFF',
        'STATUTORY HOLIDAY', 'TRAINING', 'VACATION',
    )

    def _is_nonbill(desc):
        d = desc.upper()
        return any(d.startswith(p) for p in _NON_BILL_PREFIXES)

    # Excel customer names that are non-billable and the Q360 description they map to
    _EXCEL_NONBILL_MAP = {
        'SICK':             'PERSONAL TIME-OFF',
        'PTO':              'PERSONAL TIME-OFF',
        'VACATION':         'VACATION',
        'HOLIDAY':          'STATUTORY HOLIDAY',
        'STAT':             'STATUTORY HOLIDAY',
        'ADMINISTRATION':   'ADMINISTRATION',
        'ADMIN':            'ADMINISTRATION',
        'ON-CALL':          'ON-CALL',
        'ONCALL':           'ON-CALL',
        'TRAINING':         'TRAINING',
    }

    try:
        # Load prefs first so both row-matching and dropdown use the same task IDs
        from app.db import get_db as _get_db
        _db = _get_db()
        _pref_rows = _db.execute('SELECT username, description, task_id FROM user_project_pref').fetchall()
        _prefs = {(row['username'], row['description']): row['task_id'] for row in _pref_rows}

        def _chosen_task(username, billable_by_desc_local):
            """For each description, pick the preferred task if saved, else first candidate. Returns desc→(rzk,item)."""
            result = {}
            for desc, candidates in billable_by_desc_local.items():
                pref_id = _prefs.get((username, desc))
                if pref_id:
                    match = next(((rzk, item) for rzk, item in candidates if rzk == pref_id), None)
                    result[desc] = match or candidates[0]
                else:
                    result[desc] = candidates[0]
            return result

        def _build_billable_by_desc(source):
            bd = {}
            for rzk, item in source.items():
                if Q360Service._is_admin_project(item.get('title', '')):
                    continue
                desc = item.get('description', '').strip()
                if desc and not _is_nonbill(desc):
                    bd.setdefault(desc, []).append((rzk, item))
            return bd

        # Guess resq_zoom_key for each row by text-matching Excel Project → Q360 description.
        # Row matching uses the same chosen task as the dropdown to guarantee consistency.
        for u in grouped:
            live = live_by_user.get(u['username'], {})
            if not live:
                continue
            bbd = _build_billable_by_desc(live)
            chosen_map = _chosen_task(u['username'], bbd)  # desc → (rzk, item)
            only_billable = list(chosen_map.values())  # for single-project fallback

            for r in u['rows']:
                qid = r.get('q360id', '')
                if qid and qid in live:
                    desc = live[qid].get('description', '').strip()
                    if desc:
                        r['customer'] = desc
                    r['needs_attention'] = False
                    continue

                # Non-billable customer → find matching Q360 non-billable task
                cust_upper = r.get('customer', '').upper().strip()
                target_nb_desc = _EXCEL_NONBILL_MAP.get(cust_upper)
                if target_nb_desc:
                    best_nb_rzk, best_nb_id = None, -1
                    for rzk, item in live.items():
                        desc = item.get('description', '').upper()
                        if desc.startswith(target_nb_desc.upper()):
                            if int(rzk) > best_nb_id:
                                best_nb_id, best_nb_rzk = int(rzk), rzk
                    if best_nb_rzk:
                        item = live[best_nb_rzk]
                        r['q360id']          = best_nb_rzk
                        r['customer']        = item.get('description', '').strip()
                        r['category']        = item.get('category', '') or DEFAULT_CATEGORY
                        r['needs_attention'] = False
                        r['project_guessed'] = True
                        r['suggested']       = True
                    continue

                # Text-match against billable descriptions, then use chosen task for that desc
                best_desc, best_score = None, 0.0
                for desc in chosen_map:
                    score = _match_score(r.get('customer', ''), r.get('project', ''), desc)
                    if score > best_score:
                        best_score, best_desc = score, desc

                if best_desc and best_score > 0:
                    rzk, item = chosen_map[best_desc]
                elif len(only_billable) == 1:
                    # Only one billable project — auto-assign without needing a text match
                    rzk, item = only_billable[0]
                else:
                    continue

                r['q360id']          = rzk
                r['customer']        = item.get('description', '').strip()
                r['category']        = item.get('category', '') or DEFAULT_CATEGORY
                r['needs_attention'] = False
                r['project_guessed'] = True
                r['suggested']       = True

        # Build a pool of all non-admin project items from the whole team.
        # Q360 projectscheduleno (resq_zoom_key) is global — the same task ID
        # appears in every team member's list, so cross-user matching is safe.
        all_pool = {}
        for _u in grouped:
            for rzk, item in live_by_user.get(_u['username'], {}).items():
                if not Q360Service._is_admin_project(item.get('title', '')) and rzk not in all_pool:
                    all_pool[rzk] = item

        # Second pass: for users whose Q360 fetch returned nothing, match their
        # Excel customer name against the pooled team project list.
        for u in grouped:
            if live_by_user.get(u['username'], {}):
                continue  # already handled above
            for r in u['rows']:
                customer = r.get('customer', '').strip()
                project  = r.get('project', '').strip()
                if not customer and not project:
                    continue
                # Skip non-billable rows — don't match them to billable pool projects
                if _EXCEL_NONBILL_MAP.get(customer.upper()):
                    continue
                best_rzk, best_score = None, 0.0
                for rzk, item in all_pool.items():
                    if _is_nonbill(item.get('description', '')):
                        continue
                    score = _match_score(customer, project, item.get('description', ''))
                    if score > best_score:
                        best_score, best_rzk = score, rzk
                if best_rzk and best_score > 0:
                    item = all_pool[best_rzk]
                    r['q360id']          = best_rzk
                    r['customer']        = item.get('description', '').strip()
                    r['category']        = item.get('category', '') or DEFAULT_CATEGORY
                    r['needs_attention'] = False
                    r['project_guessed'] = True
                    r['suggested']       = True

        # Build dropdown options. Users with no live data fall back to the pool
        # so the project selector is still populated for manual overrides.
        def _make_combo(rzk, item):
            desc = item.get('description', '').strip()
            if not desc:
                desc = (item.get('projecttitle') or item.get('q360_projecttitle') or '').strip()
            if not desc:
                desc = rzk  # last resort: use the ID itself
            cat = item.get('category', DEFAULT_CATEGORY) or DEFAULT_CATEGORY
            return {
                'customer': desc, 'project': '', 'q360id': rzk, 'category': cat,
                'q360_customerno':   item.get('customerno', ''),
                'q360_description':  desc,
                'q360_invoiceno':    item.get('invoiceno', ''),
                'q360_opporno':      item.get('opporno', ''),
                'q360_projectno':    item.get('projectno', ''),
                'q360_company':      item.get('company', ''),
                'q360_sitecity':     item.get('sitecity', ''),
                'q360_projecttitle': item.get('projecttitle', ''),
            }

        # Load saved task-ID preferences per user from DB
        from app.db import get_db as _get_db
        _db = _get_db()
        _pref_rows = _db.execute('SELECT username, description, task_id FROM user_project_pref').fetchall()
        _prefs = {}  # (username, description) → task_id
        for row in _pref_rows:
            _prefs[(row['username'], row['description'])] = row['task_id']

        user_projects = {}
        for u in grouped:
            live = live_by_user.get(u['username'], {})
            source = live if live else all_pool
            billable_by_desc = {}  # description → list of (rzk, item)
            nonbill_best = {}      # normalized description → (rzk, item) keeping highest numeric rzk
            for rzk, item in source.items():
                if Q360Service._is_admin_project(item.get('title', '')):
                    continue
                desc = item.get('description', '').strip()
                if not desc:
                    continue
                if _is_nonbill(desc):
                    key = desc.upper()
                    cur = nonbill_best.get(key)
                    if cur is None or int(rzk) > int(cur[0]):
                        nonbill_best[key] = (rzk, item)
                else:
                    billable_by_desc.setdefault(desc, []).append((rzk, item))

            # For each billable description, pick preferred task_id if saved; else first candidate
            billable_combos = []
            for desc, candidates in billable_by_desc.items():
                pref_id = _prefs.get((u['username'], desc))
                if pref_id:
                    match = next(((rzk, item) for rzk, item in candidates if rzk == pref_id), None)
                    chosen = match or candidates[0]
                else:
                    chosen = candidates[0]
                billable_combos.append(_make_combo(chosen[0], chosen[1]))

            combos = billable_combos + [_make_combo(rzk, item) for rzk, item in nonbill_best.values()]
            user_projects[u['username']] = sorted(combos, key=lambda p: p['customer'].lower())

    except Exception:
        # Fall back: build user_projects from Excel data so the table still renders
        user_projects = {}
        for u in grouped:
            combos = []
            seen_keys = set()
            for r in u['rows']:
                if not r['needs_attention'] and r['customer']:
                    key = (r['customer'], r['project'], r['q360id'])
                    if key not in seen_keys:
                        seen_keys.add(key)
                        combos.append({'customer': r['customer'], 'project': r['project'],
                                       'q360id': r['q360id'], 'category': r['category']})
            user_projects[u['username']] = sorted(combos, key=lambda p: p['customer'].lower())

    all_weeks = sorted(set(r['week_num'] for u in grouped for r in u['rows']))
    week_mondays = {w: _week_monday(w).strftime('%Y-%m-%d') for w in all_weeks}
    # Compute current Excel week number using the same logic as _week_monday
    _today = datetime.today().date()
    _year = _today.year
    _week0 = datetime.fromisocalendar(_year - 1, 52, 1).date()
    current_week_num = (_today - _week0).days // 7
    return render_template('bulk/_table.html', grouped=grouped, categories=CATEGORIES,
                           default_category=DEFAULT_CATEGORY, active_days=active_days,
                           user_projects=user_projects,
                           all_weeks=all_weeks, week_mondays=week_mondays,
                           current_week_num=current_week_num,
                           date_filter=date_filter)


@bp.route('/review', methods=['POST'])
@login_required
def review():
    entries_json = request.form.get('entries', '[]')
    try:
        entries = json.loads(entries_json)
    except Exception:
        return '<div class="alert alert-danger">Invalid submission data.</div>'

    # Collect date ranges per user so we can check what's already in the system
    from collections import defaultdict
    user_dates = defaultdict(set)
    for e in entries:
        if e.get('q360id'):
            for d in e.get('days', []):
                if d.get('date'):
                    user_dates[e['username']].add(d['date'])

    # Fetch existing timebills for each user/range
    existing = {}  # "username|date" -> total hours already logged
    api_errors = {}  # username -> error message
    svc = _svc()
    for username, dates in user_dates.items():
        if not dates:
            continue
        # Pre-mark all dates as clear (0h); overwritten if timebills are found
        for d in dates:
            existing[f"{username}|{d}"] = 0.0
        try:
            uid15 = username[:15]  # Q360 TIMEBILL.USERID is VARCHAR(15)
            result = svc.get_hours(uid15, min(dates), max(dates))
            for tb in result.get('userReport', []):
                date = tb['startDate'][:10]
                key = f"{username}|{date}"
                existing[key] = existing.get(key, 0.0) + float(tb.get('hours', 0))
        except Exception as e:
            # Remove pre-initialized keys so they show as "? unknown" on failure
            for d in dates:
                existing.pop(f"{username}|{d}", None)
            api_errors[username] = str(e)

    # Compute expected start/end times for display in the review table.
    # Start each day at 08:00 + existing hours already logged, then stack entries.
    review_stacks = defaultdict(list)
    for e in entries:
        if not e.get('q360id'):
            continue
        for d in e.get('days', []):
            if d.get('date') and float(d.get('hours', 0)) > 0:
                review_stacks[(e['username'], d['date'])].append(
                    (int(e.get('row_num', 0)), d)
                )
    for (username, date), stack in review_stacks.items():
        stack.sort(key=lambda x: x[0])
        existing_h = existing.get(f"{username}|{date}", 0.0)
        current = datetime.strptime(date + ' 08:00', '%Y-%m-%d %H:%M') + timedelta(hours=existing_h)
        for _, d in stack:
            d['start_time'] = current.strftime('%H:%M')
            end_dt = current + timedelta(hours=float(d['hours']))
            d['end_time'] = end_dt.strftime('%H:%M')
            current = end_dt

    return render_template('bulk/_review.html',
                           entries=entries,
                           existing=existing,
                           api_errors=api_errors,
                           entries_json=entries_json)


@bp.route('/submit', methods=['POST'])
@login_required
def submit():
    from flask import Response, stream_with_context
    entries_json = request.form.get('entries', '[]')
    try:
        entries = json.loads(entries_json)
    except Exception:
        return '<div class="alert alert-danger">Invalid submission data.</div>'

    return Response(
        stream_with_context(_submit_stream(entries)),
        mimetype='text/event-stream',
        headers={'X-Accel-Buffering': 'no', 'Cache-Control': 'no-cache'},
    )


def _submit_stream(entries):
    import json as _json
    import traceback as _tb
    from collections import defaultdict

    def _evt(obj):
        return f'data: {_json.dumps(obj)}\n\n'

    try:
        svc = _svc()
    except Exception as e:
        yield _evt({'type': 'error', 'html':
                    f'<div class="alert alert-danger">Q360 login failed: {e}</div>'})
        return

    # Count submittable day-entries so the frontend knows the total
    submittable = [
        (e, d) for e in entries
        if not e.get('needs_attention') and e.get('q360id')
        for d in e.get('days', [])
        if float(d.get('hours', 0)) > 0 and d.get('date')
    ]
    total = len(submittable)
    yield _evt({'type': 'progress', 'phase': 'check', 'done': 0, 'total': total,
                'msg': f'Checking existing hours\u2026'})

    # Pre-fetch projects and existing hours for each unique user (one call each, not per entry)
    user_dates = defaultdict(set)
    usernames_needed = set()
    for e, d in submittable:
        user_dates[e['username']].add(d['date'])
        usernames_needed.add(e['username'])

    user_projects_cache = {}
    for username in usernames_needed:
        try:
            user_projects_cache[username] = svc.get_projects(username)
        except Exception:
            user_projects_cache[username] = {}

    existing = {}
    for username, dates in user_dates.items():
        try:
            uid15 = username[:15]  # Q360 TIMEBILL.USERID is VARCHAR(15)
            result = svc.get_hours(uid15, min(dates), max(dates))
            for tb in result.get('userReport', []):
                key = f"{username}|{tb['startDate'][:10]}"
                existing[key] = existing.get(key, 0.0) + float(tb.get('hours', 0))
        except Exception:
            pass

    # Recalculate time slots server-side based on row_num ordering
    day_stacks = defaultdict(list)
    for e in entries:
        if e.get('needs_attention') or not e.get('q360id'):
            continue
        for d in e.get('days', []):
            if not d.get('date'):
                continue
            hours = float(d.get('hours', 0))
            key = f"{e['username']}|{d['date']}"
            existing_h = existing.get(key, 0.0)
            if hours > 0 and existing_h + hours <= 24.0:
                day_stacks[(e['username'], d['date'])].append((int(e['row_num']), e, d))

    for (username, date), stack in day_stacks.items():
        stack.sort(key=lambda x: x[0])
        existing_h = existing.get(f"{username}|{date}", 0.0)
        current = datetime.strptime(date + ' 08:00', '%Y-%m-%d %H:%M') + timedelta(hours=existing_h)
        for _, _, d in stack:
            d['start_time'] = current.strftime('%H:%M')
            end = current + timedelta(hours=float(d['hours']))
            d['end_time'] = end.strftime('%H:%M')
            current = end

    results = []
    done = 0

    for e in entries:
        if e.get('needs_attention') or not e.get('q360id'):
            continue
        for d in e.get('days', []):
            if not d.get('date'):
                continue
            hours = float(d.get('hours', 0))
            if hours <= 0:
                continue
            key = f"{e['username']}|{d['date']}"
            existing_h = existing.get(key, 0.0)
            if existing_h + hours > 24.0:
                results.append({'employee': e.get('username', ''), 'day': d.get('day', ''),
                                'hours': hours, 'success': None,
                                'error': f'Skipped \u2014 {existing_h}h already logged, adding {hours}h would exceed 24h'})
                done += 1
                yield _evt({'type': 'progress', 'phase': 'submit', 'done': done, 'total': total,
                            'msg': f'Skipped \u2014 {e.get("username", "")} {d.get("day", "")} {d.get("date", "")}',
                            'user': e.get('username', '')})
                continue
            try:
                date_str = d['date']
                start_time = d.get('start_time', '08:00')
                end_time = d.get('end_time', '16:00')
                sd = datetime.strptime(f"{date_str} {start_time}", '%Y-%m-%d %H:%M')
                ed = datetime.strptime(f"{date_str} {end_time}", '%Y-%m-%d %H:%M')
                if ed <= sd:
                    ed = sd + timedelta(hours=hours)
                s = f"{date_str}%20{start_time[:2]}%3A{start_time[3:5]}%3A00.000"
                e_str = f"{date_str}%20{end_time[:2]}%3A{end_time[3:5]}%3A00.000"
                delta = ed - sd
                logtime = f"{delta.total_seconds() / 3600:.2f}"
                task_data = user_projects_cache.get(e['username'], {}).get(e['q360id'])
                svc.submit_hours(
                    e['q360id'], s, e_str, logtime, e.get('comment') or None,
                    e.get('company', 'CONNEX TELECOMMUNICATIONS INC.'),
                    e['username'], e['category'],
                    task_data=task_data,
                )
                results.append({'employee': e.get('username', ''), 'day': d['day'],
                                'hours': hours, 'success': True, 'error': None})
            except Exception as ex:
                results.append({'employee': e.get('username', ''), 'day': d.get('day', ''),
                                'hours': hours, 'success': False, 'error': str(ex)})
            done += 1
            yield _evt({'type': 'progress', 'phase': 'submit', 'done': done, 'total': total,
                        'msg': f'Submitting \u2014 {e.get("username", "")} {d.get("day", "")} {d.get("date", "")}',
                        'user': e.get('username', '')})

    # Save username+description→task_id preferences for every successfully submitted entry
    try:
        from app.db import get_db as _get_db
        _db = _get_db()
        seen_prefs = set()
        for e in entries:
            if e.get('needs_attention') or not e.get('q360id'):
                continue
            username = e.get('username', '').strip()
            desc = e.get('customer', '').strip()
            task_id = e['q360id']
            if not username or not desc:
                continue
            key = (username, desc)
            if key in seen_prefs:
                continue
            seen_prefs.add(key)
            _db.execute(
                "INSERT INTO user_project_pref (username, description, task_id, updated_at) "
                "VALUES (?, ?, ?, datetime('now','localtime')) "
                "ON CONFLICT(username, description) DO UPDATE SET task_id=excluded.task_id, updated_at=excluded.updated_at",
                (username, desc, task_id)
            )
        _db.commit()
    except Exception:
        pass

    try:
        html = render_template('bulk/_result.html', results=results)
        yield _evt({'type': 'complete', 'html': html})
    except Exception as ex:
        yield _evt({'type': 'error', 'html':
                    f'<div class="alert alert-danger">Render error: {ex}</div>'})


@bp.route('/overtime/parse', methods=['POST'])
@login_required
def overtime_parse():
    """Parse overtime Excel and return a formatted output Excel file."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    file = request.files.get('file')
    if not file or not file.filename:
        return ({'error': 'No file uploaded'}, 400)

    try:
        # Read ALL sheets and concatenate — input file may have one sheet per person
        raw = pd.read_excel(file, dtype=str, sheet_name=None)
        frames = []
        for _sheet_df in raw.values():
            _sheet_df.columns = [str(c).strip() for c in _sheet_df.columns]
            frames.append(_sheet_df)
        df = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    except Exception as ex:
        return ({'error': f'Could not read Excel file: {ex}'}, 400)

    if df.empty:
        return ({'error': 'No data found in file'}, 400)

    # Normalise column names (strip whitespace, case-insensitive match)
    df.columns = [str(c).strip() for c in df.columns]
    col_map = {c.lower(): c for c in df.columns}
    cols = list(df.columns)  # ordered list for index lookups

    def _col(name):
        return col_map.get(name.lower())

    # Heuristic detection for columns whose headers are wrong/data-values:
    # - Date column:   header that can be parsed as a date (e.g. "04/03/2026")
    # - Client column: column immediately to the right of the date column
    # - Hours column:  header that is a plain number (e.g. "4")
    _detected = {}
    for i, c in enumerate(cols):
        cl = c.lower()
        # Already found by name
        if cl in ('date', 'customer', 'client', 'hours'):
            continue
        # Date-like header?
        if 'date' not in _detected:
            try:
                pd.to_datetime(c, dayfirst=False)
                _detected['date'] = c
                if i + 1 < len(cols):
                    _detected['customer'] = cols[i + 1]
                continue
            except Exception:
                pass
        # Numeric header → Hours candidate (prefer the one after ON CALL SUPPORT)
        try:
            float(c)
            if 'hours' not in _detected:
                _detected['hours'] = c
        except ValueError:
            pass

    # Override: column immediately to the right of ON CALL SUPPORT is the hours column
    on_call_col = _col('on call support')
    if on_call_col in cols:
        on_call_idx = cols.index(on_call_col)
        if on_call_idx + 1 < len(cols):
            _detected['hours'] = cols[on_call_idx + 1]

    def _col_or_detected(name):
        """Return actual column label for a logical field name."""
        found = _col(name)
        if found is not None:
            return found
        # alias: 'customer' and 'client' are the same field
        if name.lower() == 'customer':
            found = _col('client') or _detected.get('customer')
        elif name.lower() == 'client':
            found = _col('customer') or _detected.get('customer')
        else:
            found = _detected.get(name.lower())
        return found

    def _get(row, name):
        col = _col_or_detected(name)
        if col is not None and col in row.index:
            return row[col]
        # Last-resort positional fallback (original format: col 6=date,7=customer,9=hours)
        _POS = {'id':0,'start time':1,'completion time':2,'name':4,
                'date':6,'customer':7,'on call support':8,'hours':9}
        idx = _POS.get(name.lower())
        if idx is not None and idx < len(row):
            return row.iloc[idx]
        return ''

    # Date range filter using Start time column
    date_filter = request.form.get('date_filter', 'all')
    if date_filter != 'all':
        now = datetime.now()
        if date_filter == 'week':
            _f_start = now - timedelta(days=now.weekday())
            _f_end = _f_start + timedelta(days=6)
        elif date_filter == '3months':
            # Current month + 2 months prior = start of 2 months ago
            _m = now.month - 2
            _y = now.year + (_m - 1) // 12
            _m = (_m - 1) % 12 + 1
            _f_start = now.replace(year=_y, month=_m, day=1)
            _f_end = now
        elif date_filter == 'month':
            _f_start = now.replace(day=1)
            _f_end = now
        elif date_filter == 'last_month':
            _first = now.replace(day=1)
            _f_end = _first - timedelta(days=1)
            _f_start = _f_end.replace(day=1)
        elif date_filter == 'year':
            _f_start = now.replace(month=1, day=1)
            _f_end = now
        else:
            _f_start = _f_end = None

        if _f_start:
            def _parse_start(val):
                s = str(val).strip()
                for fmt in ('%Y-%m-%d %H:%M:%S', '%m/%d/%Y %H:%M:%S',
                            '%m/%d/%y %H:%M:%S', '%Y-%m-%d', '%m/%d/%Y', '%m-%d-%Y'):
                    try:
                        return datetime.strptime(s, fmt)
                    except ValueError:
                        pass
                try:
                    return pd.to_datetime(s).to_pydatetime()
                except Exception:
                    return None

            st_col = _col_or_detected('start time') or (df.columns[1] if len(df.columns) > 1 else None)
            if st_col and st_col in df.columns:
                def _in_range(row):
                    d = _parse_start(row[st_col])
                    if d is None:
                        return False
                    return _f_start.date() <= d.date() <= _f_end.date()
                df = df[df.apply(_in_range, axis=1)]

    # Load pay period lookup: date -> (pay_period, pay_week)
    db = get_db()
    pp_rows = db.execute(
        "SELECT pay_period, week1_start, week1_end, week2_start, week2_end FROM pay_period"
    ).fetchall()

    # Build fast lookup: YYYY-MM-DD -> (pay_period, week_num)
    date_to_pp = {}
    for row in pp_rows:
        pp, w1s, w1e, w2s, w2e = row
        from datetime import date as _date
        w1s_d = _date.fromisoformat(w1s)
        w1e_d = _date.fromisoformat(w1e)
        w2s_d = _date.fromisoformat(w2s)
        w2e_d = _date.fromisoformat(w2e)
        d = w1s_d
        while d <= w1e_d:
            date_to_pp[d.isoformat()] = (pp, 1)
            d += timedelta(days=1)
        d = w2s_d
        while d <= w2e_d:
            date_to_pp[d.isoformat()] = (pp, 2)
            d += timedelta(days=1)

    OUTPUT_COLS = ['ID', 'Start time', 'Completion time', 'Name', 'Date',
                   'Client', 'Work', '# of Extra Hours', 'Pay Period', 'Pay Week']

    # Group rows by person name
    from collections import defaultdict
    by_person = defaultdict(list)

    for _, row in df.iterrows():
        raw_date = str(_get(row, 'date')).strip()
        # Try to parse the date robustly
        parsed_date = None
        for fmt in ('%m/%d/%Y', '%Y-%m-%d', '%m/%d/%y', '%d/%m/%Y', '%m-%d-%Y'):
            try:
                parsed_date = datetime.strptime(raw_date, fmt).date()
                break
            except ValueError:
                continue
        if parsed_date is None:
            try:
                parsed_date = pd.to_datetime(raw_date, dayfirst=False).date()
            except Exception:
                pass

        pay_period_label = ''
        pay_week = ''
        if parsed_date:
            pp_info = date_to_pp.get(parsed_date.isoformat())
            if pp_info:
                pay_period_label, pay_week = pp_info

        # Guard against pandas NaT (which is truthy-falsy ambiguous)
        try:
            date_display = parsed_date.strftime('%-m/%-d/%Y') if parsed_date else raw_date
        except Exception:
            date_display = raw_date
            parsed_date = None

        name = str(_get(row, 'name')).strip()
        # Skip rows with no meaningful name (blank/NaN rows)
        if not name or name.lower() in ('nan', 'none', ''):
            continue
        by_person[name].append({
            'ID':               str(_get(row, 'id')).strip(),
            'Start time':       str(_get(row, 'start time')).strip(),
            'Completion time':  str(_get(row, 'completion time')).strip(),
            'Name':             name,
            'Date':             date_display,
            'Client':           str(_get(row, 'customer')).strip(),
            'Work':             str(_get(row, 'on call support')).strip(),
            '# of Extra Hours': str(_get(row, 'hours')).strip(),
            'Pay Period':       pay_period_label,
            'Pay Week':         str(pay_week),
        })

    # Persist processed records to DB
    db = get_db()
    db.execute('''CREATE TABLE IF NOT EXISTS overtime_record (
        id              INTEGER PRIMARY KEY AUTOINCREMENT,
        source_id       TEXT,
        name            TEXT NOT NULL,
        start_time      TEXT,
        completion_time TEXT,
        date            TEXT,
        client          TEXT,
        work            TEXT,
        extra_hours     REAL,
        pay_period      TEXT,
        pay_week        INTEGER,
        submitted_by    TEXT,
        parsed_at       TEXT NOT NULL DEFAULT (datetime('now','localtime'))
    )''')
    submitted_by = session.get('user_id', '')
    for person_rows in by_person.values():
        for r in person_rows:
            try:
                eh = float(r['# of Extra Hours']) if r['# of Extra Hours'] not in ('', 'nan') else None
            except (ValueError, TypeError):
                eh = None
            try:
                pw = int(r['Pay Week']) if r['Pay Week'] not in ('', 'nan') else None
            except (ValueError, TypeError):
                pw = None
            db.execute(
                '''INSERT INTO overtime_record
                   (source_id, name, start_time, completion_time, date, client, work,
                    extra_hours, pay_period, pay_week, submitted_by)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?)''',
                (r['ID'], r['Name'], r['Start time'], r['Completion time'],
                 r['Date'], r['Client'], r['Work'], eh, r['Pay Period'], pw, submitted_by)
            )
    db.commit()

    # Build output workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default sheet

    header_font = Font(bold=True)
    header_fill = PatternFill('solid', fgColor='DDEBF7')
    thin = Side(style='thin', color='BFBFBF')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for person_name, person_rows in sorted(by_person.items()):
        # Sheet names max 31 chars, strip invalid chars
        sheet_title = re.sub(r'[\\/*?:\[\]]', '', person_name)[:31]
        ws = wb.create_sheet(title=sheet_title)

        # Header row
        for col_idx, col_name in enumerate(OUTPUT_COLS, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', wrap_text=True)

        # Data rows
        for row_idx, data_row in enumerate(person_rows, 2):
            for col_idx, col_name in enumerate(OUTPUT_COLS, 1):
                val = data_row[col_name]
                # Try numeric for hours, pay week
                if col_name in ('# of Extra Hours', 'Pay Week'):
                    try:
                        val = int(val) if val != '' else val
                    except ValueError:
                        try:
                            val = float(val)
                        except ValueError:
                            pass
                cell = ws.cell(row=row_idx, column=col_idx, value=val)
                cell.border = border
                cell.alignment = Alignment(horizontal='left')

        # Auto-fit column widths
        for col_idx, col_name in enumerate(OUTPUT_COLS, 1):
            max_len = len(col_name)
            for row_idx in range(2, ws.max_row + 1):
                v = ws.cell(row=row_idx, column=col_idx).value
                if v is not None:
                    max_len = max(max_len, len(str(v)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)

    if not wb.sheetnames:
        return ({'error': 'No data rows found in file'}, 400)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from urllib.parse import quote
    fname = 'overtime_output.xlsx'
    return send_file(
        buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=fname,
    )


@bp.route('/save-username', methods=['POST'])
@login_required
def save_username():
    employee_name = request.form.get('employee_name', '').strip()
    q360_username = request.form.get('q360_username', '').strip().upper()
    if not employee_name or not q360_username:
        return ('Missing fields', 400)
    db = get_db()
    db.execute(
        'INSERT INTO username_map (employee_name, q360_username) VALUES (?, ?)'
        ' ON CONFLICT(employee_name) DO UPDATE SET q360_username=excluded.q360_username',
        (employee_name, q360_username)
    )
    db.commit()
    return ('', 204)
